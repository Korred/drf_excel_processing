from rest_framework import viewsets
from rest_framework import permissions
from rest_framework import status
from rest_framework.decorators import action
from excel_api.models import ExcelDocument
from excel_api.serializers import ExcelDocumentSerializer, ExcelSummarySerializer
from rest_framework.response import Response
from openpyxl import load_workbook
import os


class ExcelDocumentViewSet(viewsets.ModelViewSet):
    queryset = ExcelDocument.objects.all()
    serializer_class = ExcelDocumentSerializer

    # for ease of use, do not require authorization
    permission_classes = [permissions.AllowAny]

    def destroy(self, request, pk=None):
        instance = self.get_object()
        # first delete file from disc
        os.remove(instance.file.path)
        # after that remove instance
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @staticmethod
    def create_summary(document, columns):
        summary_list = []
        workbook = load_workbook(document.file.path)
        sheet = workbook[workbook.sheetnames[0]]

        if columns:
            for cell in sheet[1]:
                if cell.value in columns:

                    values = [
                        sheet[i][cell.column - 1].value
                        for i in range(2, sheet.max_row + 1)
                    ]

                    try:
                        entry = {
                            "column": cell.value,
                            "sum": sum(values),
                            "avg": round(sum(values) / len(values), 2),
                        }
                        summary_list.append(entry)
                    except TypeError:
                        # just skip this column for now
                        pass

        summary_container = {
            "file": document.file.name.split("/")[-1],
            "summary": summary_list,
        }
        return summary_container

    @action(detail=True, methods=["post"], serializer_class=ExcelSummarySerializer)
    def summary(self, request, pk=None):

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        columns = serializer.data["columns"]
        document = self.get_object()

        return Response(self.create_summary(document, columns))
